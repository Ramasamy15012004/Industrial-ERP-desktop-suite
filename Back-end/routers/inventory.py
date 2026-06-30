from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_, text, desc, func
from datetime import datetime
from pydantic import BaseModel
from typing import List, Optional, Any
import json
import csv
import io
import os
from openpyxl import load_workbook

from database import get_db
from dependencies import require_roles, ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT
from models import (
    StockMaintenance, 
    MaterialTransactionHistory, 
    MaterialMaster,
    StockTransaction,
    JobMaterialAllocation,
    Job,
    FixtureMaster,
    FixtureJobAllocation
)
from generic_utils import (
    normalize_ci_text,
    normalize_upload_column_name,
    get_purchase_upload_value,
    parse_optional_float,
    normalize_lead_days_value,
    ist_now,
    ist_date_str,
    is_generic_purchase_article,
    normalize_material_key,
    normalize_product_details,
    validate_lead_days_value,
    count_part_name_difference
)
from database_utils import (
    log_user_activity, 
    select_stock_row_by_article, 
    reallocate_pending_fixtures
)

router = APIRouter()

class Material(BaseModel):
    material_code: str
    material_name: str
    minimum_stock: int

class InwardEntry(BaseModel):
    material_code: str
    quantity: float
    transaction_type: str

class StockConfirmRow(BaseModel):
    part_name: str
    article_number: str
    make: str
    qty: float
    is_new: bool
    minimum_stock: float = 0
    lead_days: str = ""
    price: float = 0

class StockConfirmPayload(BaseModel):
    rows: List[StockConfirmRow]

class ExcelUploadPayload(BaseModel):
    rows: List[Any]

# --- Material CRUD ---

@router.post("/materials")
def create_material(material: Material, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    new_material = MaterialMaster(
        material_code=material.material_code,
        material_name=material.material_name,
        minimum_stock=material.minimum_stock
    )
    db.add(new_material)
    db.commit()
    log_user_activity(db, user, "MATERIAL_CREATED", f"Created material: {material.material_code}")
    return {"message": "Material Created Successfully"}

@router.get("/materials")
def get_materials(db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY, ROLE_AUDIT))):
    materials = db.query(MaterialMaster).all()
    return [{
        "id": m.id,
        "material_code": m.material_code,
        "material_name": m.material_name,
        "minimum_stock": m.minimum_stock
    } for m in materials]

@router.put("/materials/{id}")
def update_material(id: int, material: Material, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    existing = db.query(MaterialMaster).filter(MaterialMaster.id == id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Material not found")
    
    existing.material_code = material.material_code
    existing.material_name = material.material_name
    existing.minimum_stock = material.minimum_stock
    db.commit()
    log_user_activity(db, user, "MATERIAL_UPDATED", f"Updated material: {material.material_code}")
    return {"message": "Material Updated Successfully"}

@router.delete("/materials/{id}")
def delete_material(id: int, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    existing = db.query(MaterialMaster).filter(MaterialMaster.id == id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Material not found")
    
    db.delete(existing)
    db.commit()
    log_user_activity(db, user, "Material Delete", f"ID: {id}")
    return {"message": "Material Deleted Successfully"}

# --- Stock Maintenance ---

@router.get("/stock-maintenance")
def get_stock_maintenance(db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))):
    items = db.query(StockMaintenance).filter(
        or_(
            StockMaintenance.status == 'active',
            StockMaintenance.status == None
        )
    ).all()
    return [{
        "id": item.id,
        "part_name": item.part_name,
        "article_number": item.article_number,
        "make": item.make,
        "qty": item.qty,
        "minimum_stock": item.minimum_stock,
        "lead_days": item.lead_days,
        "price": item.price,
        "last_purchase_date": item.last_purchase_date,
        "reserved_qty": item.reserved_qty
    } for item in items]

@router.delete("/stock-maintenance/{item_id}")
def delete_stock_item(item_id: int, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    item = db.query(StockMaintenance).filter(StockMaintenance.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
        
    detail_parts = [item.part_name or f"ID {item_id}"]
    if item.article_number:
        detail_parts.append(f"Article: {item.article_number}")
    if item.make:
        detail_parts.append(f"Make: {item.make}")
    log_user_activity(db, user, "Material Delete", " | ".join(detail_parts))
    return {"message": "Item deleted successfully"}

# --- Inventory Transactions ---

@router.post("/inventory/inward")
def inward_entry(entry: InwardEntry, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    material = db.query(MaterialMaster).filter(MaterialMaster.material_code == entry.material_code).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
        
    material_name = material.material_name
    last_txn = db.query(StockTransaction).filter(StockTransaction.material_code == entry.material_code).order_by(desc(StockTransaction.id)).first()
    previous_stock = last_txn.current_stock if last_txn else 0
    
    if entry.transaction_type not in ["OPENING", "PURCHASE"]:
        raise HTTPException(status_code=400, detail="Invalid inward transaction type")
    if entry.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero")
        
    new_stock = previous_stock + entry.quantity
    now = ist_now()
    
    txn = StockTransaction(
        transaction_date=now,
        material_code=entry.material_code,
        material_name=material_name,
        quantity=entry.quantity,
        transaction_type=entry.transaction_type,
        current_stock=new_stock
    )
    db.add(txn)
    db.commit()
    
    reallocate_pending_jobs(db, entry.material_code)
    # Also check fixtures (New: Parity with reserve count requirements)
    from database_utils import reallocate_pending_fixtures
    reallocate_pending_fixtures(db, "", entry.material_code, "")
    db.commit()
    
    log_user_activity(db, user, "INVENTORY_INWARD", f"Inward material: {entry.material_code}, Qty: {entry.quantity}")
    
    return {"message": "Inward entry recorded", "current_stock": new_stock}

@router.get("/stock-transactions")
def get_stock_transactions(
    material_code: Optional[str] = None,
    transaction_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))
):
    query = db.query(StockTransaction)
    
    if material_code:
        query = query.filter(func.lower(func.trim(StockTransaction.material_code)) == func.lower(func.trim(material_code)))
    
    if transaction_type:
        query = query.filter(func.lower(func.trim(StockTransaction.transaction_type)) == func.lower(func.trim(transaction_type)))
        
    if from_date:
        query = query.filter(StockTransaction.transaction_date >= from_date)
        
    if to_date:
        query = query.filter(StockTransaction.transaction_date <= to_date)
        
    txns = query.order_by(StockTransaction.transaction_date.desc(), StockTransaction.id.desc()).limit(1000).all()
    
    return [{
        "id": t.id,
        "transaction_date": t.transaction_date,
        "material_code": t.material_code,
        "material_name": t.material_name,
        "quantity": t.quantity,
        "job_id": t.job_id,
        "transaction_type": t.transaction_type,
        "current_stock": t.current_stock
    } for t in txns]

@router.get("/material-transaction-history")
def get_material_transaction_history(
    product_details: Optional[str] = None,
    article_number: Optional[str] = None,
    bo_part_name: Optional[str] = None,
    transaction_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))
):
    query = db.query(MaterialTransactionHistory)
    
    if product_details:
        query = query.filter(func.lower(func.trim(MaterialTransactionHistory.product_details)).like(f"%{normalize_ci_text(product_details)}%"))
    
    if article_number:
        query = query.filter(func.lower(func.trim(MaterialTransactionHistory.article_number)).like(f"%{normalize_ci_text(article_number)}%"))
        
    if bo_part_name:
        query = query.filter(func.lower(func.trim(MaterialTransactionHistory.bo_part_name)).like(f"%{normalize_ci_text(bo_part_name)}%"))
        
    if transaction_type:
        query = query.filter(func.lower(func.trim(MaterialTransactionHistory.transaction_type)) == func.lower(func.trim(transaction_type)))
    
    from sqlalchemy import Date, cast
    
    if from_date:
        query = query.filter(cast(MaterialTransactionHistory.event_time, Date) >= from_date)
        
    if to_date:
        query = query.filter(cast(MaterialTransactionHistory.event_time, Date) <= to_date)
        
    history = query.order_by(MaterialTransactionHistory.event_time.desc(), MaterialTransactionHistory.id.desc()).limit(1000).all()
    
    return [{
        "id": h.id,
        "bo_no": h.bo_no,
        "product_details": h.product_details,
        "bo_part_name": h.bo_part_name,
        "article_number": h.article_number,
        "make": h.make,
        "transaction_type": h.transaction_type,
        "qty": h.qty,
        "event_time": h.event_time
    } for h in history]

@router.get("/inventory/low-stock-materials")
def get_low_stock_materials(
    make: Optional[str] = None,
    db: Session = Depends(get_db), 
    _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))
):
    query = db.query(StockMaintenance).filter(
        StockMaintenance.status == 'active',
        (StockMaintenance.qty - func.coalesce(StockMaintenance.reserved_qty, 0)) < StockMaintenance.minimum_stock,
        StockMaintenance.minimum_stock > 0
    )

    if make:
        query = query.filter(func.lower(func.trim(func.coalesce(StockMaintenance.make, ''))).like(f"%{normalize_ci_text(make)}%"))

    # Order by theoretical shortage (desc), matching SQL DESC logic
    items = query.order_by(
        (StockMaintenance.minimum_stock - (StockMaintenance.qty - func.coalesce(StockMaintenance.reserved_qty, 0))).desc()
    ).all()

    return [{
        "id": i.id,
        "part_name": i.part_name,
        "article_number": i.article_number,
        "make": i.make,
        "qty": i.qty,
        "reserved_qty": i.reserved_qty,
        "minimum_stock": i.minimum_stock,
        "lead_days": i.lead_days
    } for i in items]

@router.get("/inventory/shortage-qty")
def get_inventory_shortage_qty(
    product_details: Optional[str] = None,
    part_name: Optional[str] = None,
    article_number: Optional[str] = None,
    make: Optional[str] = None,
    db: Session = Depends(get_db), 
    _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))
):
    query = db.query(FixtureJobAllocation).join(
        FixtureMaster, 
        func.lower(func.trim(func.coalesce(FixtureMaster.product_details, ''))) == func.lower(func.trim(func.coalesce(FixtureJobAllocation.product_details, '')))
    ).filter(
        FixtureJobAllocation.shortage_qty > 0,
        func.coalesce(FixtureMaster.status, '') != 'Issued'
    )
    
    if product_details:
        query = query.filter(func.lower(func.trim(func.coalesce(FixtureJobAllocation.product_details, ''))) == normalize_product_details(product_details))
    if part_name:
        query = query.filter(func.lower(func.trim(func.coalesce(FixtureJobAllocation.bo_part_name, ''))) .like(f"%{normalize_ci_text(part_name)}%"))
    if article_number:
        query = query.filter(func.lower(func.trim(func.coalesce(FixtureJobAllocation.article_number, ''))) .like(f"%{normalize_ci_text(article_number)}%"))
    if make:
        query = query.filter(func.lower(func.trim(func.coalesce(FixtureJobAllocation.make, ''))) .like(f"%{normalize_ci_text(make)}%"))

    allocations = query.order_by(
        func.lower(func.trim(func.coalesce(FixtureJobAllocation.bo_part_name, ''))).asc(),
        func.lower(func.trim(func.coalesce(FixtureJobAllocation.article_number, ''))).asc(),
        func.lower(func.trim(func.coalesce(FixtureJobAllocation.make, ''))).asc(),
        FixtureJobAllocation.id.asc()
    ).all()
    
    results = []
    for a in allocations:
        # Get lead_days and check existence in StockMaintenance
        sm = db.query(StockMaintenance).filter(
            func.lower(func.trim(func.coalesce(StockMaintenance.part_name, ''))) == func.lower(func.trim(func.coalesce(a.bo_part_name, ''))),
            func.lower(func.trim(func.coalesce(StockMaintenance.article_number, ''))) == func.lower(func.trim(func.coalesce(a.article_number, ''))),
            StockMaintenance.status == 'active'
        ).order_by(desc(StockMaintenance.id)).first()
        
        results.append({
            "product_details": a.product_details,
            "part_name": a.bo_part_name,
            "article_number": a.article_number,
            "make": a.make,
            "required_qty": a.required_qty,
            "reserved_qty": a.reserved_qty,
            "shortage_qty": a.shortage_qty,
            "lead_days": sm.lead_days if sm else None,
            "missing_in_stock_maintenance": sm is None
        })
        
    return results

# --- Purchase Upload Logic ---

def parse_purchase_upload_file(filename: str, content: bytes) -> List[dict]:
    suffix = os.path.splitext(str(filename or "").lower())[1]
    if suffix == ".csv":
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = content.decode("latin-1")
            
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(decoded[:2048])
            delimiter = dialect.delimiter
        except Exception:
            delimiter = "\t" if "\t" in decoded else ","
        reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
        return [dict(row) for row in reader]

    if suffix == ".xlsx":
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = wb[wb.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        if not values: return []
        headers = [str(v or "").strip() for v in values[0]]
        parsed_rows = []
        for raw_row in values[1:]:
            if not raw_row: continue
            row_dict = {}
            has_value = False
            for idx, header in enumerate(headers):
                if not header: continue
                val = raw_row[idx] if idx < len(raw_row) else ""
                if val not in (None, ""): has_value = True
                row_dict[header] = "" if val is None else val
            if has_value: parsed_rows.append(row_dict)
        return parsed_rows
    return []

def build_purchase_preview_rows(db: Session, raw_rows: List[dict]):
    if not raw_rows: return []
    unique_rows = {}
    for i, row in enumerate(raw_rows, start=2):
        normalized = {normalize_upload_column_name(k): str(v or "").strip() for k, v in row.items()}
        article_number = get_purchase_upload_value(normalized, "article_number")
        qty_raw = get_purchase_upload_value(normalized, "qty")
        if not article_number or qty_raw == "": continue
        
        qty = parse_optional_float(qty_raw)
        price_raw = get_purchase_upload_value(normalized, "price")
        lead_days_raw = get_purchase_upload_value(normalized, "lead_days")
        
        if is_generic_purchase_article(article_number):
            key = normalize_material_key(get_purchase_upload_value(normalized, "part_name"), article_number, get_purchase_upload_value(normalized, "make"))
        else:
            key = normalize_ci_text(article_number)
            
        if key in unique_rows:
            unique_rows[key]["qty"] += qty
            if price_raw != "":
                unique_rows[key]["price"] = parse_optional_float(price_raw)
                unique_rows[key]["price_from_file"] = True
            if lead_days_raw != "":
                unique_rows[key]["lead_days"] = validate_lead_days_value(lead_days_raw, row_label=f"Row {i}")
        else:
            unique_rows[key] = {
                "part_name": get_purchase_upload_value(normalized, "part_name"),
                "article_number": article_number,
                "make": get_purchase_upload_value(normalized, "make"),
                "qty": qty,
                "minimum_stock": parse_optional_float(get_purchase_upload_value(normalized, "minimum_stock")),
                "lead_days": validate_lead_days_value(lead_days_raw, row_label=f"Row {i}") if lead_days_raw != "" else "",
                "price": parse_optional_float(price_raw) if price_raw != "" else 0,
                "price_from_file": price_raw != "",
            }
            
    result = []
    for row in unique_rows.values():
        existing = select_stock_row_by_article(db, row["article_number"], row["make"], row["part_name"])
        
        is_exact = False
        db_snapshot = None
        if existing:
            row_part_name = normalize_ci_text(row["part_name"])
            db_part_name = normalize_ci_text(existing.part_name)
            row_make = normalize_ci_text(row["make"])
            db_make = normalize_ci_text(existing.make)
            row_minimum_stock = float(row["minimum_stock"] if row["minimum_stock"] is not None else 0)
            db_minimum_stock = float(existing.minimum_stock if existing.minimum_stock is not None else 0)
            row_price = float(row["price"] if row["price"] is not None else 0)
            db_price = float(existing.price if existing.price is not None else 0)
            row_lead = normalize_lead_days_value(row["lead_days"])
            db_lead = normalize_lead_days_value(existing.lead_days)
            
            is_exact = (
                row_part_name == db_part_name and
                row_make == db_make and
                row_minimum_stock == db_minimum_stock and
                row_price == db_price and
                row_lead == db_lead
            )
            db_snapshot = {
                "part_name": existing.part_name or "",
                "article_number": existing.article_number or "",
                "make": existing.make or "",
                "qty": existing.qty,
                "minimum_stock": existing.minimum_stock,
                "lead_days": db_lead,
                "price": db_price
            }
            
        result.append({
            "part_name": row["part_name"],
            "article_number": row["article_number"],
            "make": row["make"],
            "qty": row["qty"],
            "is_new": not is_exact,
            "has_similar": existing is not None and not is_exact,
            "current_qty": existing.qty if existing else 0,
            "minimum_stock": row["minimum_stock"],
            "lead_days": row["lead_days"],
            "price": row["price"],
            "price_from_file": row["price_from_file"],
            "db_snapshot": db_snapshot
        })
    return result

@router.post("/upload-stock-maintenance/preview")
async def preview_stock_maintenance(file: Optional[UploadFile] = File(None), rows_json: Optional[str] = Form(None), db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    if file:
        content = await file.read()
        return build_purchase_preview_rows(db, parse_purchase_upload_file(file.filename or "", content))
    if rows_json:
        return build_purchase_preview_rows(db, json.loads(rows_json))
    raise HTTPException(status_code=400, detail="No data provided")

@router.post("/upload-stock-maintenance/preview-excel")
async def preview_stock_maintenance_excel(payload: ExcelUploadPayload, db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    if not payload.rows:
        raise HTTPException(status_code=400, detail="No data provided")
    return build_purchase_preview_rows(db, payload.rows)

@router.post("/upload-stock-maintenance/confirm")
async def confirm_stock_maintenance(payload: StockConfirmPayload, db: Session = Depends(get_db), user: dict = Depends(require_roles(ROLE_ADMIN, ROLE_INVENTORY))):
    now = ist_now()
    affected_items = set()
    inserted = 0
    updated = 0
    for row in payload.rows:
        existing = select_stock_row_by_article(db, row.article_number, row.make, row.part_name)
        if existing:
            updated += 1
            existing.qty += row.qty
            existing.minimum_stock = row.minimum_stock
            existing.price = row.price
            existing.last_purchase_date = now
            if row.lead_days: existing.lead_days = row.lead_days
        else:
            inserted += 1
            new_item = StockMaintenance(
                part_name=row.part_name, 
                article_number=row.article_number, 
                make=row.make, 
                qty=row.qty, 
                minimum_stock=row.minimum_stock, 
                price=row.price, 
                last_purchase_date=now, 
                lead_days=row.lead_days
            )
            db.add(new_item)
            
        # Record in Material Transaction History (Parity with demo_sql)
        history_entry = MaterialTransactionHistory(
            article_number=row.article_number,
            make=row.make,
            bo_part_name=row.part_name,
            transaction_type="PURCHASE ENTRY",
            qty=row.qty,
            event_time=now
        )
        db.add(history_entry)
        
        affected_items.add((row.part_name, row.article_number, row.make))
        
    total_qty = sum(float(row.qty or 0) for row in payload.rows)
    detail_parts = [
        f"Rows: {len(payload.rows)}",
        f"Inserted: {inserted}",
        f"Updated: {updated}",
        f"Qty: {total_qty:g}",
    ]
    if payload.rows:
        detail_parts.append(f"Last Article: {payload.rows[-1].article_number}")
    
    log_user_activity(db, user, "Purchase Entry", " | ".join(detail_parts))
    db.commit()

    # Reallocate shortages (New: Parity with demo_sql/app.py)
    from database_utils import reallocate_pending_fixtures
    for part_name, article_number, make in affected_items:
        reallocate_pending_fixtures(db, part_name, article_number, make)
        reallocate_pending_jobs(db, article_number)
    
    db.commit() # Final commit for reallocations
    return {"message": f"{inserted} new rows inserted, {updated} rows updated"}

# --- Helpers ---

def reallocate_pending_jobs(db: Session, material_code: str):
    last_txn = db.query(StockTransaction).filter(StockTransaction.material_code == material_code).order_by(desc(StockTransaction.id)).first()
    available_stock = last_txn.current_stock if last_txn else 0
    if available_stock <= 0: return

    allocations = db.query(JobMaterialAllocation, Job).join(Job, JobMaterialAllocation.job_id == Job.job_id).filter(JobMaterialAllocation.material_code == material_code, JobMaterialAllocation.shortage_qty > 0).order_by(Job.job_date.asc()).all()

    for alloc, job in allocations:
        if available_stock <= 0: break
        reserve_now = min(alloc.shortage_qty, available_stock)
        alloc.reserved_qty += reserve_now
        alloc.shortage_qty -= reserve_now
        available_stock -= reserve_now
        material = db.query(MaterialMaster).filter(MaterialMaster.material_code == material_code).first()
        db.add(StockTransaction(transaction_date=ist_now(), material_code=material_code, material_name=material.material_name if material else "Unknown", quantity=reserve_now, job_id=job.job_id, transaction_type="RESERVED", current_stock=available_stock))
        
        job_allocations = db.query(JobMaterialAllocation).filter(JobMaterialAllocation.job_id == job.job_id).all()
        if not remaining_shortages:
            job.status = 'Reserved'
            job.shortage_details = None
        else:
            job.status = 'Partially Reserved'
            job.shortage_details = json.dumps(remaining_shortages)

@router.get("/inventory/low-stock")
def get_low_stock_simple(db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))):
    results = db.query(MaterialMaster).all()
    low_stock = []
    for m in results:
        last_txn = db.query(StockTransaction).filter(StockTransaction.material_code == m.material_code).order_by(desc(StockTransaction.id)).first()
        current = last_txn.current_stock if last_txn else 0
        if current < m.minimum_stock:
            low_stock.append({
                "material_code": m.material_code,
                "material_name": m.material_name,
                "minimum_stock": m.minimum_stock,
                "current_stock": current
            })
    return low_stock

@router.get("/stock-maintenance/similar")
def find_similar_stock(
    part_name: str = "",
    article_number: str = "",
    make: str = "",
    db: Session = Depends(get_db), 
    _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))
):
    normalized_article = normalize_ci_text(article_number)
    normalized_part_name = normalize_ci_text(part_name)
    normalized_make = normalize_ci_text(make)

    if not normalized_article:
        return []

    candidates = db.query(StockMaintenance).filter(
        func.lower(func.trim(func.coalesce(StockMaintenance.article_number, ''))) == normalized_article,
        func.coalesce(StockMaintenance.status, 'active') == 'active'
    ).order_by(StockMaintenance.id.asc()).all()

    if not candidates:
        return []

    results = []
    for c in candidates:
        if is_generic_purchase_article(normalized_article):
            if normalized_make and normalize_ci_text(c.make) != normalized_make:
                continue
            if normalized_part_name and count_part_name_difference(c.part_name, normalized_part_name) > 1:
                continue
            match_score = count_part_name_difference(c.part_name, normalized_part_name) if normalized_part_name else 0
        else:
            make_penalty = 0 if not normalized_make or normalize_ci_text(c.make) == normalized_make else 1
            match_score = make_penalty

        results.append({
            "id": c.id,
            "part_name": c.part_name,
            "article_number": c.article_number,
            "make": c.make,
            "qty": c.qty,
            "minimum_stock": c.minimum_stock,
            "lead_days": normalize_lead_days_value(c.lead_days),
            "price": c.price if c.price is not None else 0,
            "_match_score": match_score,
        })

    results.sort(key=lambda x: (x["_match_score"], x["id"]))
    for row in results:
        del row["_match_score"]
    return results

@router.get("/inventory/total-shortage-qty")
def get_total_shortage_qty_scalar(db: Session = Depends(get_db), _: dict = Depends(require_roles(ROLE_ADMIN, ROLE_PRODUCTION, ROLE_INVENTORY, ROLE_AUDIT))):
    res = db.query(func.sum(JobMaterialAllocation.shortage_qty)).scalar() or 0
    return {"total_shortage_qty": float(res)}
